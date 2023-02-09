# pip install python-docx 
from openpyxl import load_workbook
from docx.shared import RGBColor
#import docx
from docx import Document
import sys
class edgham:
    exfilename=sys.argv[1]+".xlsx"
    wb = load_workbook(filename = exfilename)
    ws=wb.active
    cclist=list()
    mylist=list()
    i=0
    def getlecturercode(self,lname):
          rownum=self.ws.max_row
          for i in range(1,rownum+1):
            if self.ws.cell(i,2).value==lname:
                  lcode=self.ws.cell(i,1).value
          return lcode      
        
    def getlecturernames(self):
        temp_list=list()
        rownum=self.ws.max_row
        for i in range(2,rownum+1):
          temp_list.append(self.ws.cell(i,2).value)
        return temp_list
    def get_lecturer_course(self,lname):
        temp_list=list()
        mainlist=list()
        rownum=self.ws.max_row
        #mainlist.append(lname)
        for i in range(1,rownum+1):
          
            if  self.ws.cell(i,2).value==lname:
              temp_list.append(self.ws.cell(i,15).value) 
              temp_list.append(self.ws.cell(i,14).value) 
              temp_list.append(self.ws.cell(i,5).value)
              temp_list.append(self.ws.cell(i,6).value)
              
        mainlist.append(temp_list)
        temp_list.clear    
        return mainlist

    def writetoword(self,names,temp_list):
        document = Document()
        font = document.styles['Normal'].font
        font.rtl=True
        font.name="B Yekan"
        table = document.add_table(rows=1, cols=4)
        table.style.font.name="B Yekan"
        table.Alignment="center"
        row = table.add_row().cells
        i=0
        cou=0
        for k in temp_list:
          row = table.add_row().cells
          
          row[3].text=names[cou]+"("+self.getlecturercode(names[cou])+")"
          row[2].text=""
          row[1].text=""
          row[0].text=""
          row = table.add_row().cells
          
          cou+=1
          for x in k[0]:
              row[i].text=str(x)
              i+=1
              if i==4:
                    i=0
                    row = table.add_row().cells
     
        document.save('demo.docx')
    def finalcheck(self,temp_list):
          newcs=list()
          newcs=temp_list
         
         
          for x in range(0,len(temp_list)):
             for k in range(0,len(temp_list[x])):
                  for z in range(0,len(temp_list[x][k])):
                        
                        if z%4==0:
                          if temp_list[x][k].count(temp_list[x][k][z])>1:
                             newcs[x][k][z+3]=temp_list[x][k][z+3]+"**"
                              
                  
          
          return newcs
ob=edgham()
# lets get lecturer names
tp=set(ob.getlecturernames())
lnames=list(tp)
lnames.sort()
cs=list()
for mynames in lnames:
      cs.append(ob.get_lecturer_course(mynames))
newcs=ob.finalcheck(cs)
ob.writetoword(lnames,newcs)

